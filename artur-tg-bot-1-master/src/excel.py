import openpyxl
from datetime import date,datetime
from src.utils import get_today_date
from io import BytesIO


def fill_cells(naryad_dict,chat_id,naryad_num):
    file_path = './templates/naryad_template.xlsx'

    date = get_today_date()
    sheet_name = 'Лист1'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    year = date[0].year
    sheet['B2'] = naryad_dict[chat_id]['vessel_name']
    sheet['B3'] = naryad_dict[chat_id]['country']
    sheet['B4'] = naryad_dict[chat_id]['num_of_manifest']
    sheet['B5'] = naryad_dict[chat_id]['num_of_konosament']
    sheet['B6'] = naryad_dict[chat_id]['num_of_invoysya']
    sheet['B7'] = naryad_dict[chat_id]['date_of_arrival']
    sheet['C10'] = naryad_dict[chat_id]['name_of_gruz']
    sheet['D11'] = naryad_dict[chat_id]['amount_of_gas']
    sheet['D13'] = naryad_dict[chat_id]['amount_of_gas']
    sheet['C12'] = naryad_dict[chat_id]['gng_num']
    sheet['C17'] = naryad_dict[chat_id]['date_of_arrival']
    sheet['C8'] = f"Одержувач: {naryad_dict[chat_id]['reciever_name']}"
    sheet['A1'] = f'Наряд (рознарядка) № {year}-{naryad_num}'
    new_excel_file = f'Укр Наряд {date[1]}.xlsx'
    file_stream = BytesIO()

    workbook.save(file_stream)

    return new_excel_file, file_stream

def fill_cells_gen_act(gen_act: dict, chat_id: int, gen_act_num: int):
    file_path = './templates/gen_act_template.xlsx'
    date = get_today_date()

    sheet_name = 'Лист1'

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    sheet['A2'] = gen_act[chat_id]['repicient']
    sheet['B8'] = gen_act[chat_id]['vessel_name']
    sheet['E8'] = gen_act[chat_id]['country']
    sheet['I8'] = gen_act[chat_id]['arrival_date']
    sheet['F9'] = gen_act[chat_id]['download_start_date']
    sheet['K9'] = gen_act[chat_id]['download_end_date']
    sheet['C12'] = gen_act[chat_id]['cargo_name']
    sheet['C15'] = gen_act[chat_id]['cargo_name']
    sheet['K13'] = gen_act[chat_id]['konosament_weight']
    sheet['K16'] = gen_act[chat_id]['actual_weight']
    sheet['H45'] = date[1]
    sheet['G3'] = f'ГЕНЕРАЛЬНИЙ АКТ №{gen_act_num}\nGENERAL STATEMENT'

    new_excel_file = f'Ген Акт{gen_act_num} {date[1]}.xlsx'
    file_stream = BytesIO()
    workbook.save(file_stream)

    return new_excel_file, file_stream



def fill_cells_notification_act(gen_act: dict, chat_id: int, ntf_act_num: int):
    file_path = './templates/notification_act_template.xlsx'
    date = get_today_date()

    sheet_name = 'Лист1'

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    deficit = float(gen_act[chat_id]['actual_weight']) - float(gen_act[chat_id]['konosament_weight'])

    sheet['A2'] = gen_act[chat_id]['repicient']
    sheet['B8'] = gen_act[chat_id]['vessel_name']
    sheet['E8'] = gen_act[chat_id]['country']
    sheet['I8'] = gen_act[chat_id]['arrival_date']
    sheet['F9'] = gen_act[chat_id]['download_start_date']
    sheet['K9'] = gen_act[chat_id]['download_end_date']
    sheet['C13'] = gen_act[chat_id]['cargo_name']
    sheet['C15'] = gen_act[chat_id]['cargo_name']
    sheet['K13'] = gen_act[chat_id]['konosament_weight']
    sheet['K15'] = gen_act[chat_id]['actual_weight']
    sheet['H45'] = date[1]
    sheet['G3'] = f'АКТ-ПОВІДОМЛЕННЯ №{ntf_act_num}\nSTATEMENT NOTICE'
    sheet['I1'] = gen_act[chat_id]['konosam_num']
    sheet['K16'] = deficit
    new_excel_file = f'Акт Повідомлення{ntf_act_num} {date[1]}.xlsx'
    file_stream = BytesIO()
    workbook.save(file_stream)

    return new_excel_file, file_stream

